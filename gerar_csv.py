"""Gera o arquivo questoes.csv a partir da planilha Avaliacao_Humano_Juiz.xlsx.

Uso:
    python3 gerar_csv.py [planilha.xlsx] [saida.csv]

Este script so precisa ser executado de novo se a planilha original mudar.
No dia a dia, edite o questoes.csv diretamente (veja INSTRUCOES_CSV.md).
"""

import csv
import json
import sys

import openpyxl

SEPARADOR_TRECHOS = "\n---\n"
MARCA_TRUNCADO = "\n\n[TRECHO TRUNCADO NA PLANILHA ORIGINAL: completar]"


def _tenta_json(texto):
    """Decodifica uma string JSON, tolerando truncamento no final."""
    candidatos = [texto]
    base = texto.rstrip("\\")
    candidatos += [base + '"', base + '"]', base + '"]"']
    for c in candidatos:
        try:
            return json.loads(c)
        except (ValueError, TypeError):
            continue
    return None


def parse_gabarito(bruto):
    """Converte o gabarito da planilha em texto puro com trechos separados.

    Na planilha, o campo vem como uma string JSON contendo uma lista JSON,
    e cinco delas foram cortadas em 10240 caracteres pelo processo que gerou
    o arquivo. Aqui isso vira texto legivel e editavel.

    Retorna (texto, truncado).
    """
    if bruto is None:
        return "", False
    s = str(bruto).strip()
    if not s or s.lower().startswith("nao encontrou") or s.lower().startswith("não encontrou"):
        return "", False

    truncado = s.endswith("...")

    interno = _tenta_json(s) if s.startswith('"') else s
    if not isinstance(interno, str):
        interno = s

    lista = _tenta_json(interno) if interno.lstrip().startswith("[") else None
    if not isinstance(lista, list):
        lista = [interno]

    trechos = [str(t).strip() for t in lista if str(t).strip()]
    texto = SEPARADOR_TRECHOS.join(trechos)
    if truncado:
        texto += MARCA_TRUNCADO
    return texto, truncado


def main():
    entrada = sys.argv[1] if len(sys.argv) > 1 else "Avaliacao_Humano_Juiz.xlsx"
    saida = sys.argv[2] if len(sys.argv) > 2 else "questoes.csv"

    ws = openpyxl.load_workbook(entrada, data_only=True).active

    linhas = []
    for r in range(2, ws.max_row + 1):
        id_pergunta = ws.cell(r, 1).value
        pergunta = ws.cell(r, 3).value
        if not id_pergunta or not pergunta:
            continue  # o rodape da planilha traz a rubrica, nao questoes

        perfil = ws.cell(r, 2).value
        if isinstance(perfil, float) and perfil.is_integer():
            perfil = int(perfil)

        gabarito, truncado = parse_gabarito(ws.cell(r, 5).value)
        linhas.append(
            {
                "id_pergunta": str(id_pergunta).strip(),
                "perfil": "" if perfil is None else str(perfil).strip(),
                "pergunta": str(pergunta).strip(),
                "resposta": str(ws.cell(r, 4).value or "").strip(),
                "gabarito": gabarito,
            }
        )
        n_trechos = len(gabarito.split(SEPARADOR_TRECHOS)) if gabarito else 0
        print(
            "{:<8} perfil={:<3} trechos={} {}".format(
                linhas[-1]["id_pergunta"],
                linhas[-1]["perfil"],
                n_trechos,
                "TRUNCADO" if truncado else "",
            )
        )

    with open(saida, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(
            f, fieldnames=["id_pergunta", "perfil", "pergunta", "resposta", "gabarito"]
        )
        w.writeheader()
        w.writerows(linhas)

    print("\n{} questoes gravadas em {}".format(len(linhas), saida))


if __name__ == "__main__":
    main()
